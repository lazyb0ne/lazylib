import asyncio

import numpy as np
import pandas as pd
import os
import time
import openpyxl
import json

from openpyxl.reader.excel import load_workbook
from playwright.async_api import async_playwright
import xlwt
from datetime import datetime
from playwright.sync_api import sync_playwright
import re
import sys
import warnings


# 20230413 config
video_url_list = []
video_url_list_url = '../input/视频网站整理.xlsx'


def video_read_url():
    print(os.getcwd())
    # 读取Excel文件
    df = pd.read_excel(video_url_list_url)
    # 按行读取数据R
    for index, row in df.iterrows():
        # print(row['序号'], row['首页'], row['展示渠道'], row['恶意域名'])
        # print(row['column1'], row['column2'])
        video_url_list.append(row['首页'])
    print("video_read_url() count = " + str(len(video_url_list)))


def clearText(text):
    # 检查参数是否为空
    if not text:
        return False
    text = text.strip()
    if len(text) <= 3:
        return False
    return text


def video_content_by_url():
    start_time = time.time()
    # cur_time = datetime.now().strftime("%Y-%m-%d_%H:%M:%S")
    cur_time = datetime.now().strftime("%Y-%m-%d")
    # search_results_path = f'../output/VideoUrlResult{cur_time}.xlsx'
    search_results_path = f'../output/VideoUrlResult{cur_time}.xlsx'
    # openpyxl config
    # workbook = openpyxl.Workbook()
    try:
        workbook = load_workbook(search_results_path)
    except:
        workbook = openpyxl.Workbook()
        workbook.save(search_results_path)

    worksheet = workbook.active
    worksheet['A1'] = 'Web - All - Kind'
    worksheet['B1'] = 'FROM'
    worksheet['C1'] = 'Web/Mobile'
    worksheet['D1'] = 'Kind:text/external_url/internal_url/text_mobile/external_url_mobile/internal_url_mobile'
    worksheet['E1'] = 'Content'
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 40
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 30
    worksheet.column_dimensions['E'].width = 100
    worksheet.freeze_panes = 'A2'
    video_url_list_count = len(video_url_list)
    url_ok = get_web_url(search_results_path)
    print(url_ok)
    print(f"Already get: {len(url_ok)}")
    xls_row_count = 0
    for video_idx, video_url in enumerate(video_url_list):
        # video_url = 'http://www.nongkenfang.com' # 测试站点
        print(f"Current url:{video_url}")
        if is_valid_url(video_url):
            continue
        if video_idx < 296:
            continue
        is_continue = False
        try:
            if url_ok and video_url in url_ok:
                # 最后一条删掉重新跑
                if video_url == url_ok[-1]:
                    print(f"Last need to delete:   {video_url}")
                    word_in_xlsx(search_results_path, video_url)
                    print(f"Deleted OK:   {video_url}")
                else:
                    print(f"continue   {video_url}")
                    is_continue = True
                    continue
        except:
            print("error")
        if is_continue:
            continue
        try:
            page_text, internal_links, external_links = download_website_text_and_links(video_url, False)
            page_text_mobile, internal_links_mobile, external_links_mobile = download_website_text_and_links(video_url, True)
        except:
            continue
        list_all = [page_text, internal_links, external_links, page_text_mobile, internal_links_mobile,
                    external_links_mobile]
        list_web_mobile = ['web', 'web', 'web', 'mobile', 'mobile', 'mobile']
        list_kink = ['text', 'internal_links', 'external_links', 'text_mobi', 'internal_links_mobi',
                     'external_links_mobi']

        for i in range(len(list_all)):
            url_cost = time.time()
            kind_index = 0
            for item in list_all[i]:
                kind_cost = time.time()
                item = clearText(item)
                if item is False:
                    continue
                kind_index += 1
                first_data = str(video_idx) + " - " + str(xls_row_count) + " - " + str(kind_index)
                # print(first_data)
                if list_kink[i] in ['internal_links', 'internal_links_mobi'] and video_url not in item:
                    item = video_url + item
                try:
                    worksheet.append([first_data, video_url, list_web_mobile[i], list_kink[i], item])
                    xls_row_count += 1
                    show_info(video_idx, xls_row_count, kind_index, start_time, video_url_list_count, url_cost,
                              kind_cost)
                except Exception:
                    pass

        if video_idx > 300:
            break
        workbook.save(search_results_path)


def show_info(video_idx, xls_row_count, kind_index, start_time, video_url_list_count, url_cost, kind_cost):
    t_all = time.time() - start_time
    t_url = time.time() - url_cost
    t_kind = time.time() - kind_cost
    print("{} / {} kind: {} cost(sec) ALL/URL/Kind:{:.2f} / {:.2f} / {:.2f}".format(video_url_list_count, video_idx, kind_index, t_all, t_url, t_kind))
    # print(f"video:{video_idx}")


# 下载某个url所有的标签、文本
def download_website_text_and_links(url, isPhone=False, headless=True):
    with sync_playwright() as playwright:
        if isPhone:
            iphone = playwright.devices['iPhone 12 Pro Max']
            browser = playwright.webkit.launch(headless=headless)
            context = browser.new_context(**iphone, locale='zh-CN')
            page = context.new_page()
        else:
            browser = playwright.chromium.launch(headless=headless)
            # browser = playwright.chromium.launch()
            context = browser.new_context()
            page = context.new_page()
        page.goto(url)
        # await page.wait_for_selector('body')
        page_text = page.inner_text('body')  # 获取页面的文本内容
        page_text = page_text.split("\n")
        unique_list = list(dict.fromkeys(page_text))

        page_links = [link.get_attribute('href') for link in page.query_selector_all('a')]  # 获取页面中所有链接的 href 属性值
        browser.close()

        internal_links = []  # 保存内链
        external_links = []  # 保存外链
        for link in page_links:
            if link.startswith('http'):  # 判断链接是否以 http 开头，即为绝对路径
                if url in link:  # 如果链接中包含基准 URL，即为内链
                    internal_links.append(link)
                else:  # 否则为外链
                    external_links.append(link)
            else:  # 如果链接不是以 http 开头，即为相对路径，也视为内链
                internal_links.append(link)

    return unique_list, internal_links, external_links


def lazy():
    cur_time = datetime.now().strftime("%Y-%m-%d_%H:%M:%S")
    search_results_path = f'../output/VideoUrlResult{cur_time}.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet['A1'] = 'Web - All - Kind'
    worksheet['B1'] = 'FROM'
    worksheet['C1'] = 'Web/Mobile'
    worksheet['D1'] = 'Kind:text/external_url/internal_url/text_mobile/external_url_mobile/internal_url_mobile'
    worksheet['E1'] = 'Content'
    # 写入数据到多个单元格
    data = [
        ['Name', 'Age', 'City'],
        ['Alice', 25, 'New York'],
        ['Bob', 30, 'Los Angeles'],
        ['Charlie', 35, 'Chicago']
    ]
    for row in data:
        worksheet.append(row)
        # 保存 Excel 文件
    workbook.save('example.xlsx')


def is_valid_url(url):
    try:
        pattern = re.compile(
            r'^(?:http|ftp)s?://'  # scheme
            r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?))'  # domain
            r'(?:/?|[/?]\S+)$', re.IGNORECASE)
        result = bool(re.match(pattern, url))
    except:
        result = False
    return result


def get_last_row_data(xlsx_file):
    workbook = openpyxl.load_workbook(xlsx_file)
    sheet = workbook.active
    last_row = sheet.max_row
    last_row_values = []
    for column in sheet.iter_cols(min_row=last_row, max_row=last_row, values_only=True):
        last_row_values = column
    last_row_values = [cell.value for cell in sheet[last_row]]
    return last_row_values


def get_web_url(xlsx_file):
    workbook = openpyxl.load_workbook(xlsx_file)
    sheet = workbook.active
    web_url_ok = []
    for row in sheet.iter_rows():
        web_url_ok.append(row[1].value)
    return np.unique(web_url_ok) | None

def word_in_xlsx(file, str_to_delete):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    word_found = False
    rows_to_delete = []

    # Iterate through the rows and check if the word is present in the row data
    for row in range(1, sheet.max_row + 1):
        print("--------")
        print(row)
        # row_data = [cell.value for cell in sheet[row]]  # Get the values of cells in the row
        row_data = [sheet[row][1].value]  # Get the values of cells in the row
        print(row_data)

        if str_to_delete in row_data:
            rows_to_delete.append(row)  # Add the row number to the list of rows to delete

    # Delete the rows in reverse order to avoid issues with row deletion
    idx = 1
    for row in reversed(rows_to_delete):
        sheet.delete_rows(row)  # Delete the row
        print(f"{idx} Deleted: {str_to_delete}")
        idx +=1

    # Save the updated Excel file
    workbook.save(file)  # Replace 'file_updated.xlsx' with the desired name for the updated file

def test():
    file = "../output/VideoUrlResult230424.xlsx"
    # web_list = get_last_row_data(file)
    web_list = get_web_url(file)
    print(web_list)
    # word_in_xlsx(file, "https://fua.cc/")
    # print()
    # for idx, row in enumerate([11,22,33]):
    #     print(f"Deleted: {idx}")


if __name__ == '__main__':
    warnings.filterwarnings("ignore", category=DeprecationWarning)
    video_read_url()
    video_content_by_url()
    # lazy()
    # test()

